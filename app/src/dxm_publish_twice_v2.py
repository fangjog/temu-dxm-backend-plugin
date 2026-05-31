from __future__ import annotations

import datetime
import hashlib
import json
import os
import re
import sys
import time
from concurrent.futures import Future, ThreadPoolExecutor
from typing import Any

from . import dxm_publish_twice_flow as legacy
from .dianxiaomi_pages import fill_product_title, select_origin_country_and_province
from .publish_pages import (
    choose_package_image_with_dimension_priority,
    extract_dimensions_from_product_images,
    fill_variant_dimensions_and_weight_from_images,
)
from .text_ai import optimize_product_title
from .utils import PROJECT_ROOT, take_screenshot


DRAFT_PRODUCTS_URL = "https://www.dianxiaomi.com/web/temu/choiceTemuList/draft"
PUBLISHING_PRODUCTS_URL = "https://www.dianxiaomi.com/web/temu/choiceTemuList/offline?dxmOfflineState=publishing"
ONLINE_PRODUCTS_URL = "https://www.dianxiaomi.com/web/temu/choiceTemuList/online"
PUBLISH_FAIL_URL = "https://www.dianxiaomi.com/web/temu/choiceTemuList/offline?dxmOfflineState=publishFail"
COLLECT_BOX_URL = DRAFT_PRODUCTS_URL
PENDING_PRODUCTS_URL = "https://www.dianxiaomi.com/web/temu/choiceTemuList/offline"
DEFAULT_PRODUCT_COUNT = 4
TASK_VERIFY_TARGETS = [
    {"label": "online", "name": "在线产品", "url": ONLINE_PRODUCTS_URL, "screenshot": "verify_online_products"},
    {"label": "publishing", "name": "发布中产品", "url": PUBLISHING_PRODUCTS_URL, "screenshot": "verify_publishing_tasks"},
    {"label": "publishFail", "name": "发布失败产品", "url": PUBLISH_FAIL_URL, "screenshot": "verify_publish_failed_tasks"},
]
_TITLE_EXECUTOR = ThreadPoolExecutor(max_workers=2, thread_name_prefix="dxm-title")

FORBIDDEN_TITLE_TERMS = [
    "100% natural sand",
    "construction sand",
    "children toy",
    "toy for kids",
    "for children",
    "for kids",
    "natural sand",
    "crushed stone",
    "volcanic sand",
    "stone powder",
    "mineral sand",
    "quartz sand",
    "silica sand",
    "desert sand",
    "river sand",
    "beach sand",
    "real sand",
    "rock sand",
    "raw mineral",
    "children",
    "toddler",
    "infant",
    "mineral",
    "quartz",
    "silica",
    "volcanic",
    "child",
    "girls",
    "boys",
    "kids",
    "baby",
    "girl",
    "boy",
    "kid",
    "sand",
]


def _log(logger: Any | None, step: str, status: str, message: str, page: Any | None = None, **extra: Any) -> None:
    try:
        legacy._log(logger, step, status, message, page=page, **extra)
    except Exception:
        pass


def _ensure_utf8_console() -> None:
    for stream in (getattr(sys, "stdout", None), getattr(sys, "stderr", None)):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def _clean_forbidden_title(title: str) -> dict[str, Any]:
    cleaned = str(title or "").strip()
    hit_terms: list[str] = []
    for term in sorted(FORBIDDEN_TITLE_TERMS, key=len, reverse=True):
        pattern = r"(?<![A-Za-z0-9])" + re.escape(term).replace(r"\ ", r"\s+") + r"(?![A-Za-z0-9])"
        if re.search(pattern, cleaned, flags=re.IGNORECASE):
            hit_terms.append(term)
            cleaned = re.sub(pattern, " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*[-,/|]+\s*", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_,./|")
    if not cleaned:
        cleaned = "Home Decor Accessory"
    return {"hit": bool(hit_terms), "terms": hit_terms, "clean_title": cleaned[:180].strip()}


def _patch_legacy_short_title_cleaner() -> None:
    if not hasattr(legacy, "_v2_original_shorten_product_title"):
        legacy._v2_original_shorten_product_title = legacy.shorten_product_title

    def wrapped(page: Any, logger: Any | None = None) -> dict[str, Any]:
        result = legacy._v2_original_shorten_product_title(page, logger=logger)
        if result.get("status") != "ok":
            return result
        title = result.get("new_title") or legacy._safe_read_title(page)
        cleaned = _clean_forbidden_title(title)
        if cleaned["hit"] or cleaned["clean_title"] != title:
            fill_product_title(page, cleaned["clean_title"], logger=logger)
            page.wait_for_timeout(500)
            result["new_title"] = cleaned["clean_title"]
            _log(logger, "second_title_forbidden_clean", "ok", f"Cleaned forbidden short-title terms: {cleaned['terms']}", page=page)
        result["forbidden_hit"] = bool(cleaned["hit"])
        result["cleaned_forbidden_terms"] = cleaned["terms"]
        return result

    legacy.shorten_product_title = wrapped


def _apply_title_clean_on_page(page: Any, context: dict[str, Any], prefix: str, logger: Any | None = None) -> None:
    title = legacy._safe_read_title(page)
    cleaned = _clean_forbidden_title(title)
    context[f"{prefix}_title_forbidden_hit"] = cleaned["hit"]
    context[f"{prefix}_cleaned_forbidden_terms"] = cleaned["terms"]
    context[f"{prefix}_cleaned_title"] = cleaned["clean_title"]
    if cleaned["hit"] or cleaned["clean_title"] != title:
        fill_product_title(page, cleaned["clean_title"], logger=logger)
        page.wait_for_timeout(500)
        _log(logger, f"{prefix}_title_forbidden_clean", "ok", f"Cleaned forbidden title terms: {cleaned['terms']}", page=page)


def _start_long_title_generation(page: Any, context: dict[str, Any], prefix: str, original_title: str, logger: Any | None = None) -> None:
    title = str(original_title or "").strip()
    if not title:
        return
    future_key = f"_{prefix}_title_future"
    if context.get(future_key):
        return
    context[future_key] = _TITLE_EXECUTOR.submit(optimize_product_title, title)
    context[f"{prefix}_title_ai_started_at"] = datetime.datetime.now().isoformat()
    _log(
        logger,
        f"{prefix}_title_generation_started",
        "start",
        "Started AI title rewrite as soon as the edit page was opened.",
        page=page,
    )


def _extract_source_products(page: Any, product_count: int, logger: Any | None = None, target_title: str = "") -> list[dict[str, Any]]:
    if _is_probably_edit_page(page):
        title = legacy._safe_read_title(page)
        return [{
            "source_index": 1,
            "title": title,
            "sku": legacy._read_first_sku_from_edit_page(page),
            "price": "",
            "row_text_preview": title,
            "current_edit_page": True,
            "edit_url": page.url,
        }]
    rows: list[dict[str, Any]] = []
    for attempt in range(4):
        rows = page.evaluate(
            """(limit) => {
                const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
                document.querySelectorAll('[data-dxm-v2-row-index]').forEach((el) => el.removeAttribute('data-dxm-v2-row-index'));
                const selectors = ['.vxe-body--row', 'tr.vxe-body--row', 'tbody tr', 'tr.ant-table-row', '.ant-table-row', '.el-table__row'];
                let rows = [];
                for (const selector of selectors) {
                    rows = Array.from(document.querySelectorAll(selector)).filter(visible).filter((row) => {
                        const text = textOf(row);
                        const rect = row.getBoundingClientRect();
                        if (rect.y < 80 || rect.height < 20 || text.length < 40) return false;
                        if (!/编辑|发布|更多|移入待发布/.test(text)) return false;
                        if (/商品信息|操作|全选|店铺账号|搜索类型|搜索内容|排序类型/.test(text) && text.length < 180) return false;
                        return true;
                    });
                    if (rows.length) break;
                }
                rows = rows.sort((a, b) => a.getBoundingClientRect().y - b.getBoundingClientRect().y).slice(0, Math.max(limit, 1));
                return rows.map((row, index) => {
                    row.setAttribute('data-dxm-v2-row-index', String(index + 1));
                    const text = textOf(row);
                    const titleEl = row.querySelector('.white-space, [class*="white-space"]');
                    const lines = text.split(/\\s{2,}|\\n|\\r/).map((line) => line.trim()).filter(Boolean);
                    let title = titleEl ? textOf(titleEl) : '';
                    if (!title) {
                        title = lines.filter((line) => line.length >= 12 && !/编辑|修改|删除|更多|操作|价格|库存|状态|店铺|时间|SKU|CNY|USD|￥|\\$|创建|更新|移入待发布|发布/i.test(line))
                            .sort((a, b) => b.length - a.length)[0] || '';
                    }
                    if (!title) {
                        const cleaned = text.replace(/^(来源|亚马逊\\([^)]*\\)|速卖通|SHEIN)\\s+/i, '');
                        title = cleaned.split(/「|--|CNY|创建:/)[0].trim();
                    }
                    const skuParts = [];
                    for (const match of text.matchAll(/\\b[0-9]{8,}\\b/g)) skuParts.push(match[0]);
                    const sku = skuParts.slice(0, 5).join(' ');
                    const price = ((text.match(/(?:USD|CNY|￥|\\$)\\s*([0-9]+(?:\\.[0-9]+)?)/i) || [])[1] || '');
                    const img = Array.from(row.querySelectorAll('img')).filter(visible)[0];
                    return {
                        source_index: index + 1,
                        title: title.slice(0, 240),
                        sku,
                        price,
                        image_src: img ? (img.currentSrc || img.src || '') : '',
                        row_text_preview: text.slice(0, 800)
                    };
                });
            }""",
            max(product_count, 30) if str(target_title or "").strip() else product_count,
        )
        rows = rows if isinstance(rows, list) else []
        if rows:
            break
        page.wait_for_timeout(700)
    wanted = str(target_title or "").strip().lower()
    if wanted and rows:
        wanted_tokens = _tokens(wanted)
        best: dict[str, Any] | None = None
        best_score = 0
        for row in rows:
            haystack = " ".join([str(row.get("title", "")), str(row.get("row_text_preview", ""))]).lower()
            score = 0
            if wanted[:45] in haystack:
                score += 100
            overlap = len(wanted_tokens & _tokens(haystack))
            score += overlap
            if score > best_score:
                best = row
                best_score = score
        if best and best_score >= max(6, min(12, len(wanted_tokens) // 2)):
            best["target_title_matched"] = True
            best["target_title"] = target_title
            rows = [best]
            _log(logger, "extract_source_products", "ok", f"Matched target source product by title, score={best_score}.", page=page, extra={"target_title": target_title, "source": best})
            return rows[:product_count]
        if rows:
            rows[0]["source_title_not_found_used_first_row"] = True
            rows[0]["target_title"] = target_title
            _log(logger, "extract_source_products", "warning", "Target title not found in current draft rows; using first row.", page=page, extra={"target_title": target_title})
            return rows[:product_count]
    _log(logger, "extract_source_products", "ok", f"Extracted {len(rows)} source product rows.", page=page)
    return rows[:product_count]


def _is_probably_edit_page(page: Any) -> bool:
    try:
        url = (page.url or "").lower()
    except Exception:
        url = ""
    if "dianxiaomi.com" in url and ("/edit" in url or "quoteedit" in url):
        return True
    try:
        return bool(legacy._is_edit_page(page))
    except Exception:
        return False


def _trim_extra_browser_pages(page: Any, max_pages: int = 3, logger: Any | None = None) -> None:
    try:
        pages = list(page.context.pages)
    except Exception:
        return
    if len(pages) <= max_pages:
        return
    keep = {id(page)}
    for candidate in pages:
        if len(keep) >= max_pages:
            break
        if candidate is page:
            continue
        try:
            url = (candidate.url or "").lower()
        except Exception:
            url = ""
        if "dianxiaomi.com" in url and ("choicetumulist/draft" in url or "dxmofflinestate=publishing" in url):
            keep.add(id(candidate))
    for candidate in pages:
        if id(candidate) in keep:
            continue
        try:
            candidate.close()
        except Exception:
            pass
    _log(logger, "trim_browser_pages", "ok", f"Trimmed browser pages from {len(pages)} to <= {max_pages}.", page=page)


def _prefer_source_list_page(page: Any) -> Any:
    try:
        pages = list(page.context.pages)
    except Exception:
        return page
    # This flow must always start from the Temu collected/draft list. Do not
    # reuse publishing, failure, online, Yunqi, or Temu front pages as source.
    target = page
    for candidate in pages:
        try:
            url = (candidate.url or "").lower()
        except Exception:
            continue
        if "dianxiaomi.com" in url:
            target = candidate
            if "choicetumulist/draft" in url:
                break
    try:
        target.bring_to_front()
    except Exception:
        pass
    try:
        target.goto(DRAFT_PRODUCTS_URL, wait_until="domcontentloaded", timeout=12000)
        legacy._wait_ready(target)
        target.wait_for_timeout(900)
    except Exception as exc:
        _log(None, "open_draft_source_list", "warning", f"Could not open draft source list: {exc}", page=target)
    return target


def _mark_source_row(page: Any, source: dict[str, Any], fallback_index: int) -> dict[str, Any]:
    return page.evaluate(
        """({source, fallbackIndex}) => {
            const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
            const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
            document.querySelectorAll('[data-dxm-v2-target-edit],[data-dxm-v2-target-more]').forEach((el) => {
                el.removeAttribute('data-dxm-v2-target-edit');
                el.removeAttribute('data-dxm-v2-target-more');
            });
            const selectors = ['tr.ant-table-row', '.ant-table-row', '.vxe-body--row', '.el-table__row', 'tbody tr', '[class*="table"] [class*="row"]', '[class*="list"] [class*="item"]'];
            let rows = [];
            for (const selector of selectors) {
                rows = Array.from(document.querySelectorAll(selector)).filter(visible).filter((row) => {
                    const text = textOf(row);
                    const rect = row.getBoundingClientRect();
                    if (rect.y < 80 || rect.height < 20 || text.length < 12) return false;
                    if (/鍟嗗搧淇℃伅|鎿嶄綔|鍏ㄩ€墊搴楅摵|浠锋牸|搴撳瓨|SKU/.test(text) && text.length < 120) return false;
                    return true;
                });
                if (rows.length) break;
            }
            rows = rows.sort((a, b) => a.getBoundingClientRect().y - b.getBoundingClientRect().y);
            const title = String(source.title || '').trim();
            const sku = String(source.sku || '').trim();
            const preview = String(source.row_text_preview || '').trim();
            const desiredIndex = Number(source.source_index || fallbackIndex || 1);
            let row = null;
            if (!row && title) row = rows.find((candidate) => textOf(candidate).includes(title.slice(0, Math.min(80, title.length))));
            if (!row && sku) row = rows.find((candidate) => textOf(candidate).includes(sku));
            if (!row && preview) row = rows.find((candidate) => textOf(candidate).includes(preview.slice(0, 80)));
            if (!row && desiredIndex >= 1 && rows[desiredIndex - 1]) row = rows[desiredIndex - 1];
            if (!row) row = rows[Math.max(0, fallbackIndex - 1)] || rows[0] || null;
            if (!row) return {ok: false, message: 'No source row found.'};
            const actionRegex = /^(编辑|修改|查看|查看\\/?编辑|同步后编辑|Edit|View|Modify)$/i;
            const createRegex = /^(创建产品|创建新产品|复制为.*|Create|Copy)$/i;
            const moreRegex = /^(更多|操作|\\.\\.\\.|more)$/i;
            const actions = Array.from(row.querySelectorAll('button, a, span, div[role="button"]')).filter(visible);
            const rowRect = row.getBoundingClientRect();
            const sameLineActions = Array.from(document.querySelectorAll('button, a, span, div[role="button"]')).filter(visible).filter((el) => {
                const rect = el.getBoundingClientRect();
                const cy = rect.y + rect.height / 2;
                return rect.x > Math.max(900, rowRect.right - 260) && cy >= rowRect.y - 8 && cy <= rowRect.bottom + 8;
            });
            const allActions = actions.concat(sameLineActions);
            let edit = sameLineActions.find((el) => actionRegex.test(textOf(el)) && !/删除|移除|批量|发布/.test(textOf(el)));
            if (!edit) edit = sameLineActions.find((el) => createRegex.test(textOf(el)) && !/删除|移除|批量|发布/.test(textOf(el)));
            if (!edit) edit = actions.find((el) => actionRegex.test(textOf(el)) && !/删除|移除|批量|发布/.test(textOf(el)));
            let more = sameLineActions.find((el) => moreRegex.test(textOf(el)) || /dropdown|more|ellipsis|down/i.test(String(el.className || '')));
            if (!more) more = actions.find((el) => moreRegex.test(textOf(el)) || /dropdown|more|ellipsis|down/i.test(String(el.className || '')));
            if (edit) edit.setAttribute('data-dxm-v2-target-edit', '1');
            if (more) more.setAttribute('data-dxm-v2-target-more', '1');
            return {ok: true, has_direct_edit: !!edit, has_more: !!more, row_text_preview: textOf(row).slice(0, 500)};
        }""",
        {"source": source, "fallbackIndex": fallback_index},
    )


def _click_text(page: Any, texts: list[str], timeout_ms: int = 2500) -> bool:
    for text in texts:
        try:
            loc = page.get_by_text(text, exact=False).last
            loc.wait_for(state="visible", timeout=timeout_ms)
            loc.scroll_into_view_if_needed(timeout=timeout_ms)
            loc.click(timeout=timeout_ms)
            return True
        except Exception:
            continue
    return False


def _click_visible_dropdown_item(page: Any, texts: list[str]) -> bool:
    try:
        return bool(
            page.evaluate(
                """(texts) => {
                    const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                    const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
                    const containers = Array.from(document.querySelectorAll('.ant-dropdown, .el-dropdown-menu, .dropdown-menu, [role="menu"], .ant-select-dropdown')).filter(visible);
                    const candidates = [];
                    for (const container of containers) {
                        candidates.push(...Array.from(container.querySelectorAll('li, a, button, div, span')).filter(visible));
                    }
                    for (const el of candidates) {
                        const text = textOf(el);
                        if (!text || /删除|移除|批量/.test(text)) continue;
                        if (texts.some((item) => text.includes(item))) {
                            el.scrollIntoView({block: 'center', inline: 'nearest'});
                            for (const type of ['mouseover', 'mouseenter', 'mousedown', 'mouseup', 'click']) {
                                el.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
                            }
                            return true;
                        }
                    }
                    return false;
                }""",
                texts,
            )
        )
    except Exception:
        return False


def _confirm_copy_store_modal_if_present(page: Any, logger: Any | None = None) -> bool:
    try:
        confirmed = bool(
            page.evaluate(
                """() => {
                    const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                    const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
                    const modals = Array.from(document.querySelectorAll('.ant-modal, .el-dialog, [role="dialog"], .modal, .ant-popover')).filter(visible);
                    const modal = modals.find((item) => textOf(item).includes('选择店铺') || textOf(item).includes('复制') || textOf(item).includes('Kyiki'));
                    if (!modal) return false;
                    const checkable = Array.from(modal.querySelectorAll('input[type="checkbox"], input[type="radio"]')).filter(visible);
                    const unchecked = checkable.find((input) => !input.checked);
                    if (unchecked) {
                        unchecked.click();
                    }
                    const buttons = Array.from(modal.querySelectorAll('button, a, span, div[role="button"]')).filter(visible);
                    const ok = buttons.find((btn) => textOf(btn) === '确定') || buttons.find((btn) => /确定|确认|OK|Confirm/i.test(textOf(btn)));
                    if (!ok) return false;
                    ok.scrollIntoView({block: 'center', inline: 'nearest'});
                    for (const type of ['mouseover', 'mouseenter', 'mousedown', 'mouseup', 'click']) {
                        ok.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
                    }
                    return true;
                }"""
            )
        )
        if confirmed:
            _log(logger, "copy_new_product_confirm", "ok", "Confirmed copy-as-new-product store modal.", page=page)
            page.wait_for_timeout(3000)
        return confirmed
    except Exception:
        return False


def _find_probable_edit_page(context: Any, preferred: Any | None = None, before_pages: list[Any] | None = None) -> Any | None:
    candidates: list[Any] = []
    pages = []
    try:
        pages = list(context.pages)
    except Exception:
        pages = []
    ordered = []
    if preferred is not None:
        ordered.append(preferred)
    if before_pages:
        ordered.extend([candidate for candidate in pages if candidate not in before_pages])
    elif not ordered:
        ordered.extend(pages)
    seen: set[int] = set()
    for candidate in ordered:
        if candidate is None:
            continue
        if before_pages and candidate in before_pages and candidate is not preferred:
            continue
        marker = id(candidate)
        if marker in seen:
            continue
        seen.add(marker)
        try:
            if _is_probably_edit_page(candidate):
                candidates.append(candidate)
        except Exception:
            continue
    return candidates[-1] if candidates else None


def _wait_for_probable_edit_page(context: Any, page: Any, before_pages: list[Any], timeout_ms: int = 20000) -> Any | None:
    deadline = time.monotonic() + (timeout_ms / 1000)
    while time.monotonic() < deadline:
        target = _find_probable_edit_page(context, preferred=page, before_pages=before_pages)
        if target is not None:
            try:
                target.bring_to_front()
            except Exception:
                pass
            try:
                legacy._wait_ready(target)
            except Exception:
                pass
            return target
        try:
            page.wait_for_timeout(500)
        except Exception:
            time.sleep(0.5)
    return None


def _click_copy_as_new_product_from_marked_row(page: Any, logger: Any | None = None) -> bool:
    try:
        opened = bool(
            page.evaluate(
                """() => {
                    const el = document.querySelector('[data-dxm-v2-target-more="1"]');
                    if (!el) return false;
                    el.scrollIntoView({block: 'center', inline: 'nearest'});
                    for (const type of ['mouseover', 'mouseenter', 'mousemove', 'mousedown', 'mouseup', 'click']) {
                        el.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
                    }
                    return true;
                }"""
            )
        )
        if not opened:
            return False
        page.wait_for_timeout(1000)
        clicked = _click_visible_dropdown_item(
            page,
            [
                "复制为“新产品”",
                "复制为\"新产品\"",
                "复制为新产品",
                "复制为",
                "创建新产品",
                "Copy",
            ],
        )
        if clicked:
            _log(logger, "copy_new_product_menu", "ok", "Clicked copy-as-new-product menu item.", page=page)
            page.wait_for_timeout(1000)
        return clicked
    except Exception:
        return False


def _open_source_edit(page: Any, source: dict[str, Any], fallback_index: int, logger: Any | None = None, state: Any | None = None) -> Any:
    if source.get("current_edit_page") and _is_probably_edit_page(page):
        return page
    edit_url = str(source.get("edit_url") or "").strip()
    if edit_url and "dianxiaomi.com" in edit_url and ("/edit" in edit_url or "quoteEdit" in edit_url):
        page.goto(edit_url, wait_until="domcontentloaded", timeout=12000)
        legacy._wait_ready(page)
        if _is_probably_edit_page(page):
            _log(logger, "open_source_product_edit", "ok", f"Opened source product by stored edit URL for #{fallback_index}.", page=page, extra={"edit_url": edit_url})
            return page
    mark = _mark_source_row(page, source, fallback_index)
    if not mark.get("ok"):
        legacy.fail_with_popup_and_screenshot(page, "open_source_product_edit", mark.get("message", "Could not locate source product row."), logger=logger, state=state, extra={"source": source})
    screenshot_path = take_screenshot(page, f"dxm_source_{fallback_index}_before_edit")
    _log(logger, "open_source_product_edit", "start", f"Opening source product #{fallback_index} edit page.", page=page, screenshot_path=screenshot_path, extra={"source": source, "mark": mark})
    context = page.context
    before_pages = list(context.pages)
    clicked = False
    try:
        clicked = bool(page.evaluate(
            """() => {
                const el = document.querySelector('[data-dxm-v2-target-edit="1"]');
                if (!el) return false;
                el.scrollIntoView({block: 'center', inline: 'center'});
                for (const type of ['mouseover', 'mouseenter', 'mousemove', 'mousedown', 'mouseup', 'click']) {
                    el.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
                }
                return true;
            }"""
        ))
    except Exception:
        clicked = False
    if not clicked:
        try:
            target = page.locator('[data-dxm-v2-target-edit="1"]').first
            target.scroll_into_view_if_needed(timeout=3000)
            target.click(timeout=5000)
            clicked = True
        except Exception:
            clicked = False
    if not clicked:
        try:
            clicked = _click_copy_as_new_product_from_marked_row(page, logger=logger)
            if not clicked:
                more = page.locator('[data-dxm-v2-target-more="1"]').first
                more.scroll_into_view_if_needed(timeout=3000)
                try:
                    more.hover(timeout=3000)
                except Exception:
                    pass
                try:
                    more.click(timeout=5000)
                except Exception:
                    pass
                page.wait_for_timeout(1200)
                clicked = _click_text(
                    page,
                    [
                        "编辑",
                        "修改",
                        "查看/编辑",
                        "同步后编辑",
                        "创建产品",
                        "创建新产品",
                        "复制为“新产品”",
                        "复制为\"新产品\"",
                        "复制为新产品",
                        "复制为",
                        "Edit",
                        "Modify",
                        "Create",
                        "Copy",
                    ],
                    timeout_ms=3000,
                )
        except Exception:
            clicked = False
    if not clicked:
        legacy.fail_with_popup_and_screenshot(page, "open_source_product_edit", "Could not click the source product edit action.", logger=logger, state=state, extra={"source": source, "mark": mark})
    page.wait_for_timeout(1000)
    for _ in range(3):
        if _confirm_copy_store_modal_if_present(page, logger=logger):
            break
        page.wait_for_timeout(800)
    target_page = _wait_for_probable_edit_page(context, page, before_pages, timeout_ms=12000) or page
    target_page.wait_for_timeout(600)
    if not _is_probably_edit_page(target_page):
        extra = {"source": source, "mark": mark}
        try:
            extra["open_pages"] = [candidate.url for candidate in context.pages]
        except Exception:
            pass
        legacy.fail_with_popup_and_screenshot(target_page, "open_source_product_edit", "Clicked edit/copy action but did not enter a Dianxiaomi Temu edit page.", logger=logger, state=state, extra=extra)
    return target_page


def _return_to_source_list(page: Any, source_list_url: str, logger: Any | None = None) -> Any:
    if not source_list_url:
        return page
    try:
        page.goto(source_list_url, wait_until="domcontentloaded", timeout=12000)
        legacy._wait_ready(page)
        page.wait_for_timeout(900)
        _log(logger, "return_source_list", "ok", "Returned to original Dianxiaomi source list.", page=page)
    except Exception as exc:
        _log(logger, "return_source_list", "failed", f"Could not return to source list: {exc}", page=page)
    return page


def _short_hash(value: str, length: int = 10) -> str:
    return hashlib.sha1(str(value or "").encode("utf-8", errors="ignore")).hexdigest()[:length]


def _make_task_marker(source: dict[str, Any], source_index: int, edit_id: str = "") -> str:
    if edit_id:
        return f"S{source_index}-{edit_id}"
    raw = "|".join([
        str(source_index),
        str(source.get("title") or ""),
        str(source.get("sku") or ""),
        str(source.get("row_text_preview") or "")[:300],
    ])
    return f"S{source_index}-{_short_hash(raw, 10)}"


def _sku_suffix_values(result: dict[str, Any]) -> tuple[str, str, str]:
    if not isinstance(result, dict):
        return "", "", ""
    suffix = str(result.get("suffix") or "")
    items = result.get("items", [])
    old_values: list[str] = []
    new_values: list[str] = []
    if isinstance(items, list):
        for item in items:
            if not isinstance(item, dict):
                continue
            old_values.append(str(item.get("old") or item.get("old_sku") or ""))
            new_values.append(str(item.get("new") or item.get("new_sku") or ""))
    return suffix, "; ".join([v for v in old_values if v]), "; ".join([v for v in new_values if v])


def _build_publish_tasks(context: dict[str, Any]) -> list[dict[str, Any]]:
    first_suffix, first_sku_old, first_sku_new = _sku_suffix_values((context.get("first_publish_result") or {}).get("sku_suffix", {}))
    second_suffix, second_sku_old, second_sku_new = _sku_suffix_values((context.get("second_publish_result") or {}).get("sku_suffix", {}))
    return [
        {
            "source_product_index": context.get("source_product_index", ""),
            "publish_round": 1,
            "task_marker": context.get("task_marker", ""),
            "edit_url": context.get("first_edit_url", ""),
            "edit_id": context.get("first_edit_id", ""),
            "is_quote_edit": False,
            "title_before": context.get("first_title_before", ""),
            "title_clean": context.get("first_title_after") or context.get("first_cleaned_title", ""),
            "title_forbidden_hit": bool(context.get("first_title_forbidden_hit")),
            "cleaned_forbidden_terms": context.get("first_cleaned_forbidden_terms") or [],
            "image_result": context.get("first_image_result", {}),
            "sku_old": first_sku_old,
            "sku_new": first_sku_new,
            "sku_suffix": first_suffix,
            "category_old": "",
            "category_new": "",
            "submit_status": context.get("first_publish_status", ""),
            "publish_time": context.get("first_publish_time", ""),
            "publishing_record_count": 0,
            "publishing_match_text": "",
            "publishing_verify_status": "not_run",
            "failure_reason": "",
        },
        {
            "source_product_index": context.get("source_product_index", ""),
            "publish_round": 2,
            "task_marker": context.get("task_marker", ""),
            "edit_url": context.get("second_edit_url", ""),
            "edit_id": context.get("second_edit_id", ""),
            "is_quote_edit": "quoteedit" in str(context.get("second_edit_url", "")).lower(),
            "title_before": context.get("second_title_before", ""),
            "title_clean": context.get("second_title_after", ""),
            "title_forbidden_hit": bool(context.get("second_title_forbidden_hit")),
            "cleaned_forbidden_terms": context.get("second_cleaned_forbidden_terms") or [],
            "image_result": context.get("second_image_result", {}),
            "sku_old": second_sku_old,
            "sku_new": second_sku_new,
            "sku_suffix": second_suffix,
            "category_old": (context.get("category_result") or {}).get("old_category", "") if isinstance(context.get("category_result"), dict) else "",
            "category_new": (context.get("category_result") or {}).get("new_category", "") if isinstance(context.get("category_result"), dict) else "",
            "submit_status": context.get("second_publish_status", ""),
            "publish_time": context.get("second_publish_time", ""),
            "publishing_record_count": 0,
            "publishing_match_text": "",
            "publishing_verify_status": "not_run",
            "failure_reason": "",
        },
    ]


def _publish_source_twice(page: Any, source: dict[str, Any], source_index: int, config: dict[str, Any], logger: Any | None = None, state: Any | None = None) -> tuple[Any, dict[str, Any]]:
    edit_page = _open_source_edit(page, source, source_index, logger=logger, state=state)
    context = legacy.build_context_from_first_row(edit_page, logger=logger)
    first_edit_id = legacy._extract_edit_context_id(edit_page.url)
    first_title_before = context.get("first_title_before") or source.get("title") or ""
    context.update({
        "source_product_index": source_index,
        "source_list_title": source.get("title") or context.get("source_list_title", ""),
        "source_list_sku": source.get("sku") or context.get("source_list_sku", ""),
        "source_list_price": source.get("price", ""),
        "source_row_text_preview": source.get("row_text_preview", ""),
        "source_title_not_found_used_first_row": bool(source.get("source_title_not_found_used_first_row")),
        "target_title": source.get("target_title", ""),
        "first_edit_url": edit_page.url,
        "first_edit_id": first_edit_id,
        "first_title_before": first_title_before,
        "run_started_at": datetime.datetime.now().isoformat(),
        "task_marker": _make_task_marker(source, source_index, first_edit_id),
        "sku_run_suffix": f"{source_index}{int(time.time() * 1000) % 100000}",
    })
    _start_long_title_generation(edit_page, context, "first", context.get("first_title_before", ""), logger=logger)
    first = legacy.run_first_publish_edit(edit_page, config, context=context, logger=logger, state=state)
    context["first_publish_result"] = first
    context["first_image_result"] = first.get("image_shuffle", {})
    context["first_price_result"] = first.get("sku_price_increase", {})
    context["first_title_after"] = first.get("title") or context.get("first_cleaned_title") or legacy._safe_read_title(edit_page)
    context["first_title_forbidden_hit"] = bool(context.get("first_title_forbidden_hit") or first.get("title_forbidden_hit"))
    context["first_cleaned_forbidden_terms"] = list(dict.fromkeys((context.get("first_cleaned_forbidden_terms") or []) + (first.get("cleaned_forbidden_terms") or [])))
    if first.get("status") in {"manual_required", "error", "failed"}:
        context["failure_reason"] = f"first_publish_status_{first.get('status')}"
        context["publish_tasks"] = _build_publish_tasks(context)
        return edit_page, context
    elif first.get("status") not in legacy.PUBLISH_SUCCESS_STATUSES:
        _log(logger, "first_publish_status_pending_verify", "warning", f"First publish returned {first.get('status')}; continuing to second publish and final task verification.", page=edit_page)
    try:
        second_page = legacy.click_second_publish_entry(edit_page, logger=logger, state=state)
        legacy.assert_second_edit_is_new_product(second_page, context, logger=logger, state=state)
        context["second_entry_mode"] = "continue_publish"
    except legacy.DxmTwiceFlowError as exc:
        _log(
            logger,
            "second_publish_entry_fallback",
            "warning",
            f"Second publish entry unavailable ({exc.step}: {exc.message}); reopening the same source product from draft for round 2.",
            page=edit_page,
            screenshot_path=exc.screenshot_path,
        )
        list_page = _return_to_source_list(edit_page, DRAFT_PRODUCTS_URL, logger=logger)
        second_page = _open_source_edit(list_page, source, source_index, logger=logger, state=state)
        context["second_entry_mode"] = "draft_reopen"
    context["second_edit_url"] = second_page.url
    context["second_edit_id"] = legacy._extract_edit_context_id(second_page.url)
    context["second_title_before"] = legacy._safe_read_title(second_page)
    second = legacy.run_second_publish_edit(second_page, config, context=context, logger=logger, state=state)
    context["second_publish_result"] = second
    context["second_image_result"] = second.get("image_shuffle", {})
    context["category_result"] = second.get("category", {})
    context["second_title_after"] = (second.get("title") or {}).get("new_title") or context.get("second_title_after") or legacy._safe_read_title(second_page)
    context["second_title_forbidden_hit"] = bool((second.get("title") or {}).get("forbidden_hit") or context.get("second_before_short_title_forbidden_hit"))
    context["second_cleaned_forbidden_terms"] = list(dict.fromkeys((context.get("second_before_short_cleaned_forbidden_terms") or []) + ((second.get("title") or {}).get("cleaned_forbidden_terms") or [])))
    if second.get("status") not in legacy.PUBLISH_SUCCESS_STATUSES:
        context["failure_reason"] = f"second_publish_status_{second.get('status')}"
    context["publish_tasks"] = _build_publish_tasks(context)
    return second_page, context


def _extract_visible_rows(page: Any, max_rows: int = 120) -> list[dict[str, Any]]:
    try:
        rows = page.evaluate(
            """(maxRows) => {
                const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
                const selectors = ['tr.ant-table-row', '.ant-table-row', '.vxe-body--row', '.el-table__row', 'tbody tr', '[class*="table"] [class*="row"]', '[class*="list"] [class*="item"]'];
                let rows = [];
                for (const selector of selectors) {
                    rows = Array.from(document.querySelectorAll(selector)).filter(visible).filter((row) => {
                        const text = textOf(row);
                        const rect = row.getBoundingClientRect();
                        if (rect.y < 80 || rect.height < 18 || text.length < 12) return false;
                        if (/商品信息|操作|全选|店铺|价格|库存|SKU/.test(text) && text.length < 120) return false;
                        return true;
                    });
                    if (rows.length) break;
                }
                rows = rows.sort((a, b) => a.getBoundingClientRect().y - b.getBoundingClientRect().y).slice(0, maxRows);
                return rows.map((row, index) => {
                    const text = textOf(row);
                    const links = Array.from(row.querySelectorAll('a[href]')).map((a) => ({text: textOf(a).slice(0, 200), href: a.href || ''})).filter((link) => link.href && !/^javascript:/i.test(link.href));
                    const ids = [];
                    for (const match of text.matchAll(/(?:ID|商品ID|产品ID|goodsId|itemId)[:：]?\\s*([A-Za-z0-9_-]+)/gi)) ids.push(match[1]);
                    for (const link of links) for (const match of link.href.matchAll(/(?:id|productId|goodsId|itemId)=([A-Za-z0-9_-]+)/gi)) ids.push(match[1]);
                    const skus = [];
                    for (const match of text.matchAll(/(?:SKU|货号)[:：]?\\s*([A-Za-z0-9_-]+)/gi)) skus.push(match[1]);
                    const titleEl = row.querySelector('.white-space, [class*="white-space"]');
                    const lines = text.split(/\\s{2,}|\\n|\\r/).map((line) => line.trim()).filter(Boolean);
                    const title = ((titleEl ? textOf(titleEl) : '') || lines.filter((line) => line.length >= 8 && !/编辑|修改|删除|更多|操作|价格|库存|状态|店铺|时间|SKU|CNY|USD|￥|\\$/i.test(line)).sort((a, b) => b.length - a.length)[0]
                        || lines.filter((line) => line.length >= 8).sort((a, b) => b.length - a.length)[0]
                        || text.slice(0, 160));
                    return {row_index: index + 1, title: title.slice(0, 240), text_preview: text.slice(0, 1200), links, ids: Array.from(new Set(ids)).slice(0, 10), skus: Array.from(new Set(skus)).slice(0, 10), page_url: location.href};
                });
            }""",
            max_rows,
        )
        return rows if isinstance(rows, list) else []
    except Exception:
        return []


def _record_text(record: dict[str, Any]) -> str:
    links = " ".join(str(item.get("text", "")) + " " + str(item.get("href", "")) for item in record.get("links", []) if isinstance(item, dict))
    return " ".join([str(record.get("title", "")), str(record.get("text_preview", "")), links])


def _publish_result_for_validation_page(label: str) -> str:
    if label == "online":
        return "success"
    if label == "publishing":
        return "publishing"
    if label == "publishFail":
        return "failed"
    return "not_found"


def _extract_publish_failure_reason(record: dict[str, Any]) -> str:
    text = re.sub(r"\s+", " ", _record_text(record)).strip()
    if not text:
        return ""

    markers = ["失败原因", "失败详情", "错误原因", "解决方案", "报错", "接口报错", "api access", "access_token"]
    positions = [text.find(marker) for marker in markers if text.find(marker) >= 0]
    if not positions:
        error_like = re.findall(r"[^。；;]*(?:失败|报错|错误|error|failed|access_token|api access)[^。；;]*[。；;]?", text, flags=re.IGNORECASE)
        return " ".join(item.strip() for item in error_like if item.strip())[:2000]

    reason = text[min(positions):].strip()
    reason = re.split(r"\s+(?:SKU|CNY|USD|创建[:：]|更新[:：]|编辑\s+发布|Temu\s+https?://)", reason, maxsplit=1)[0].strip()
    reason = re.sub(r"\s+(?:编辑|发布|更多|查看|同步|复制|删除)(?:\s+(?:编辑|发布|更多|查看|同步|复制|删除))*\s*$", "", reason)
    return reason[:2000].strip()


def _sku_values(value: str) -> list[str]:
    values: list[str] = []
    for item in re.split(r"[;,\s]+", str(value or "")):
        item = item.strip()
        if not item or item in values:
            continue
        values.append(item)
    return values


def _record_has_exact_sku(record_text: str, sku: str) -> bool:
    value = str(sku or "").strip()
    if not value:
        return False
    return bool(re.search(rf"(?<![A-Za-z0-9_-]){re.escape(value)}(?![A-Za-z0-9_-])", record_text, flags=re.IGNORECASE))


def _tokens(value: str) -> set[str]:
    stop = {"for", "and", "the", "with", "set", "pcs", "piece", "pieces", "new", "home", "temu", "kyiki", "cny", "usd", "jit", "mode"}
    return {token for token in re.findall(r"[A-Za-z0-9]+", str(value or "").lower()) if len(token) >= 3 and token not in stop}


def _record_matches_context(record: dict[str, Any], context: dict[str, Any]) -> bool:
    text = _record_text(record)
    text_lower = text.lower()
    candidates = [
        context.get("second_title_after", ""),
        context.get("first_title_after", ""),
        context.get("first_cleaned_title", ""),
        context.get("second_cleaned_title", ""),
        context.get("first_title_before", ""),
        context.get("source_list_title", ""),
    ]
    for title in [str(item or "").strip() for item in candidates if str(item or "").strip()]:
        if len(title) >= 12 and title.lower() in text_lower:
            return True
        try:
            if legacy._titles_similar(text, title, threshold=0.25):
                return True
        except Exception:
            pass
        title_tokens = _tokens(title)
        if title_tokens and len(title_tokens & _tokens(text)) >= min(6, max(4, len(title_tokens) // 2)):
            return True
    sku = str(context.get("source_list_sku") or "").strip()
    return bool(sku and sku.lower() in text_lower)


def _record_is_recent_for_context(record: dict[str, Any], context: dict[str, Any]) -> bool:
    started = str(context.get("run_started_at") or "").strip()
    if not started:
        return True
    try:
        start_dt = datetime.datetime.fromisoformat(started) - datetime.timedelta(minutes=2)
    except Exception:
        return True
    text = _record_text(record)
    matches = re.findall(r"20\\d{2}-\\d{2}-\\d{2}\\s+\\d{2}:\\d{2}", text)
    if not matches:
        return False
    for value in matches:
        try:
            if datetime.datetime.strptime(value, "%Y-%m-%d %H:%M") >= start_dt:
                return True
        except Exception:
            continue
    return False


def _unique_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    unique: list[dict[str, Any]] = []
    for record in records:
        key = legacy._record_key(record)
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)
    return unique


def _detect_failure_reasons(page: Any, contexts: list[dict[str, Any]], logger: Any | None = None) -> dict[int, str]:
    reasons: dict[int, str] = {}
    try:
        page.goto(PUBLISH_FAIL_URL, wait_until="domcontentloaded", timeout=12000)
        legacy._wait_ready(page)
        page.wait_for_timeout(3000)
        records = _extract_visible_rows(page)
        for ctx in contexts:
            idx = int(ctx.get("source_product_index", 0) or 0)
            matched = [record for record in records if _record_matches_context(record, ctx)]
            reason = ""
            for record in matched:
                text = _record_text(record)
                if "access_token don't have this api access" in text or "access_token don" in text:
                    reason = "temu_authorization_failed"
                    break
            if matched and not reason:
                reason = "publish_failed_record_found"
            if reason:
                reasons[idx] = reason
    except Exception as exc:
        _log(logger, "verify_publish_failures", "warning", f"Could not inspect publish failure page: {exc}", page=page)
    return reasons


def _verify_online_products(page: Any, contexts: list[dict[str, Any]], logger: Any | None = None, state: Any | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "status": "running",
        "verify_result": "not_run",
        "product_count": len(contexts),
        "expected_online_records": len(contexts) * 2,
        "actual_online_records": 0,
        "success_product_count": 0,
        "failed_product_count": 0,
        "products": [],
        "online_screenshot_path": "",
    }
    try:
        page.goto(ONLINE_PRODUCTS_URL, wait_until="domcontentloaded", timeout=12000)
        legacy._wait_ready(page)
        page.wait_for_timeout(1000)
        _click_text(page, ["所有分类"], timeout_ms=1500)
        page.wait_for_timeout(1000)
    except Exception as exc:
        result.update({"status": "failed", "verify_result": "dual_publish_verify_failed", "failure_reason": f"online_products_page_not_opened: {exc}"})
        return result

    online_records = _extract_visible_rows(page)
    screenshot_path = legacy._safe_take_screenshot(page, "verify_online_products")
    result["online_screenshot_path"] = screenshot_path
    failure_reasons = _detect_failure_reasons(page, contexts, logger=logger)
    actual = 0
    for ctx in contexts:
        idx = int(ctx.get("source_product_index", 0) or 0)
        matched = _unique_records([
            record for record in online_records
            if _record_is_recent_for_context(record, ctx) and _record_matches_context(record, ctx)
        ])
        links: list[str] = []
        for record in matched:
            for link in record.get("links", []) if isinstance(record.get("links"), list) else []:
                href = str(link.get("href", "")) if isinstance(link, dict) else ""
                if href and href not in links:
                    links.append(href)
        count = len(matched)
        actual += count
        failure_reason = failure_reasons.get(idx, "")
        if not failure_reason:
            if count == 0:
                failure_reason = "no_online_record_found"
            elif count < 2:
                failure_reason = "only_one_online_record_found"
        success = count >= 2
        result["success_product_count"] += 1 if success else 0
        result["failed_product_count"] += 0 if success else 1
        ctx["distinct_online_records_count"] = count
        ctx["distinct_publish_records_count"] = count
        ctx["publish_links"] = links
        ctx["verify_result"] = "dual_publish_verified" if success else "dual_publish_verify_failed"
        ctx["failure_reason"] = "" if success else failure_reason
        result["products"].append({
            "source_product_index": idx,
            "source_list_title": ctx.get("source_list_title", ""),
            "first_title": ctx.get("first_title_after") or ctx.get("first_title_before", ""),
            "second_title": ctx.get("second_title_after", ""),
            "first_edit_id": ctx.get("first_edit_id", ""),
            "second_edit_id": ctx.get("second_edit_id", ""),
            "online_records_count": count,
            "publish_links_count": len(links),
            "publish_links": links,
            "failure_reason": "" if success else failure_reason,
            "success": success,
            "online_records": matched[:4],
        })
    result["actual_online_records"] = actual
    all_success = bool(contexts) and result["success_product_count"] == len(contexts) and actual >= result["expected_online_records"]
    result["status"] = "success" if all_success else "failed"
    result["verify_result"] = "dual_publish_verified" if all_success else "dual_publish_verify_failed"
    _log(logger, "verify_online_products", result["status"], f"Online verification completed: {result['verify_result']}", page=page, screenshot_path=screenshot_path, extra=result)
    if state:
        state.update(dxm_twice_online_verify=result)
    return result


def _sku_price_summary(price_result: dict[str, Any]) -> tuple[str, str]:
    items = price_result.get("items", []) if isinstance(price_result, dict) else []
    old_values: list[str] = []
    new_values: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        old_values.append(str(item.get("old_price") or item.get("old") or item.get("before") or ""))
        new_values.append(str(item.get("new_price") or item.get("new") or item.get("after") or ""))
    return "; ".join([v for v in old_values if v]), "; ".join([v for v in new_values if v])


def _image_counts(image_result: dict[str, Any]) -> tuple[Any, Any, str, str]:
    if not isinstance(image_result, dict):
        return "", "", "", ""
    before = image_result.get("selected_before") or image_result.get("before_count") or image_result.get("initial_selected_count") or image_result.get("original_count") or ""
    after = image_result.get("selected_after") or image_result.get("after_count") or image_result.get("final_selected_count") or image_result.get("final_count") or ""
    filled_to_10 = "是" if str(after).isdigit() and int(after) >= 10 else "否"
    shuffled = "是" if image_result.get("status") == "ok" else "否"
    return before, after, filled_to_10, shuffled


def _write_reports(contexts: list[dict[str, Any]], verification: dict[str, Any]) -> dict[str, str]:
    reports_dir = PROJECT_ROOT / "data" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = reports_dir / f"dxm_publish_twice_result_{stamp}.json"
    xlsx_path = reports_dir / f"dxm_publish_twice_result_{stamp}.xlsx"
    verify_products = {int(item.get("source_product_index", 0) or 0): item for item in verification.get("products", []) if isinstance(item, dict)}
    rows: list[dict[str, Any]] = []
    for ctx in contexts:
        idx = int(ctx.get("source_product_index", 0) or 0)
        verify = verify_products.get(idx, {})
        old_price, new_price = _sku_price_summary(ctx.get("first_price_result", {}))
        first_img_before, first_img_after, first_img_full, first_img_shuffle = _image_counts(ctx.get("first_image_result", {}))
        second_img_before, second_img_after, second_img_full, second_img_shuffle = _image_counts(ctx.get("second_image_result", {}))
        category = ctx.get("category_result", {}) if isinstance(ctx.get("category_result"), dict) else {}
        online_count = verify.get("online_records_count", ctx.get("distinct_online_records_count", 0))
        links = verify.get("publish_links", ctx.get("publish_links", []))
        final_success = "成功" if online_count >= 2 else "失败"
        failure_reason = verify.get("failure_reason") or ctx.get("failure_reason", "")
        rows.append({
            "源产品序号": idx,
            "发布轮次": "第一次",
            "源商品标题": ctx.get("source_list_title", ""),
            "清理后标题": ctx.get("first_title_after") or ctx.get("first_cleaned_title", ""),
            "是否命中标题禁词": "是" if ctx.get("first_title_forbidden_hit") else "否",
            "已清理禁词": ", ".join(ctx.get("first_cleaned_forbidden_terms") or []),
            "编辑页ID": ctx.get("first_edit_id", ""),
            "是否quoteEdit": "否",
            "图片原数量": first_img_before,
            "图片最终数量": first_img_after,
            "图片是否补到10张": first_img_full,
            "图片是否打乱": first_img_shuffle,
            "SKU原价": old_price,
            "SKU新价": new_price,
            "原类目": category.get("old_category", ""),
            "新类目": category.get("new_category", ""),
            "发布状态": ctx.get("first_publish_status", ""),
            "在线记录数": online_count,
            "发布链接": links[0] if links else "",
            "失败原因": failure_reason,
            "最终是否成功": final_success,
        })
        rows.append({
            "源产品序号": idx,
            "发布轮次": "第二次",
            "源商品标题": ctx.get("source_list_title", ""),
            "清理后标题": ctx.get("second_title_after", ""),
            "是否命中标题禁词": "是" if ctx.get("second_title_forbidden_hit") else "否",
            "已清理禁词": ", ".join(ctx.get("second_cleaned_forbidden_terms") or []),
            "编辑页ID": ctx.get("second_edit_id", ""),
            "是否quoteEdit": "是" if "quoteedit" in str(ctx.get("second_edit_url", "")).lower() else "否",
            "图片原数量": second_img_before,
            "图片最终数量": second_img_after,
            "图片是否补到10张": second_img_full,
            "图片是否打乱": second_img_shuffle,
            "SKU原价": "",
            "SKU新价": "",
            "原类目": category.get("old_category", ""),
            "新类目": category.get("new_category", ""),
            "发布状态": ctx.get("second_publish_status", ""),
            "在线记录数": online_count,
            "发布链接": links[1] if len(links) > 1 else (links[0] if links else ""),
            "失败原因": failure_reason,
            "最终是否成功": final_success,
        })
    payload = {"summary": verification, "products": contexts, "rows": rows}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    try:
        import pandas as pd

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df = pd.DataFrame(rows)
            df.to_excel(writer, index=False, sheet_name="双发布结果")
            ws = writer.book["双发布结果"]
            ws.freeze_panes = "A2"
            for column_cells in ws.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 48)
    except Exception:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "双发布结果"
        headers = list(rows[0].keys()) if rows else ["结果"]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
    return {"json_path": str(json_path), "xlsx_path": str(xlsx_path)}


def _task_match_score(record: dict[str, Any], context: dict[str, Any], task: dict[str, Any]) -> tuple[int, list[str]]:
    text = _record_text(record)
    text_lower = text.lower()
    score = 0
    reasons: list[str] = []

    sku_new_values = _sku_values(str(task.get("sku_new") or ""))
    if sku_new_values:
        sku_matches = [sku for sku in sku_new_values if _record_has_exact_sku(text, sku)]
        if sku_matches:
            score += 30
            reasons.append("sku_new")

    for key, points in (("sku_suffix", 5), ("edit_id", 3), ("task_marker", 2)):
        value = str(task.get(key) or context.get(key) or "").strip()
        matched = False
        if key == "sku_suffix":
            matched = bool(value and re.search(rf"{re.escape(value)}(?![A-Za-z0-9_-])", text, flags=re.IGNORECASE))
        else:
            matched = bool(value and value.lower() in text_lower)
        if matched:
            score += points
            reasons.append(key)
    title = str(task.get("title_clean") or "").strip()
    if title:
        if len(title) >= 10 and title.lower() in text_lower:
            score += 6
            reasons.append("title_exact")
        else:
            try:
                if legacy._titles_similar(text, title, threshold=0.28):
                    score += 4
                    reasons.append("title_similar")
            except Exception:
                pass
            title_tokens = _tokens(title)
            if title_tokens:
                overlap = len(title_tokens & _tokens(text))
                if overlap >= min(7, max(4, len(title_tokens) // 2)):
                    score += 3
                    reasons.append("title_tokens")
    if _record_is_recent_for_context(record, context):
        score += 2
        reasons.append("recent")
    return score, reasons


def _task_match_accepted(score: int, reasons: list[str]) -> bool:
    if "sku_new" in reasons and score >= 8:
        return True
    if "sku_suffix" in reasons and ("title_tokens" in reasons or "title_exact" in reasons) and score >= 9:
        return True
    if "title_exact" in reasons and score >= 10:
        return True
    if "searched_title" in reasons and ("title_exact" in reasons or "title_tokens" in reasons) and score >= 10:
        return True
    return False


def _search_current_list_by_text(page: Any, query: str, logger: Any | None = None) -> list[dict[str, Any]]:
    value = str(query or "").strip()
    if not value:
        return []
    try:
        ok = page.evaluate(
            """(query) => {
                const visible = (el) => !!(el.offsetWidth || el.offsetHeight || el.getClientRects().length);
                const inputs = Array.from(document.querySelectorAll('input:not([type=hidden]), textarea')).filter(visible);
                const scored = inputs.map((el, index) => {
                    const rect = el.getBoundingClientRect();
                    const hint = `${el.placeholder || ''} ${el.getAttribute('aria-label') || ''} ${el.className || ''}`;
                    let score = 0;
                    if (/搜索|标题|商品|内容|SKU|货号|keyword|search/i.test(hint)) score += 20;
                    if (rect.y < 260) score += 10;
                    if (rect.width > 120) score += 5;
                    return {el, index, score};
                }).sort((a, b) => b.score - a.score || a.el.getBoundingClientRect().y - b.el.getBoundingClientRect().y);
                const picked = scored[0] && scored[0].score > 0 ? scored[0].el : inputs[0];
                if (!picked) return false;
                picked.focus();
                picked.value = '';
                picked.dispatchEvent(new Event('input', {bubbles: true}));
                picked.value = query;
                picked.dispatchEvent(new Event('input', {bubbles: true}));
                picked.dispatchEvent(new Event('change', {bubbles: true}));
                const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
                const buttons = Array.from(document.querySelectorAll('button, a, span, div[role=button]')).filter(visible);
                const button = buttons.find((el) => /搜索|查询|Search/i.test(textOf(el))) || buttons.find((el) => /search/i.test(String(el.className || '')));
                if (button) button.click();
                else picked.dispatchEvent(new KeyboardEvent('keydown', {key: 'Enter', code: 'Enter', bubbles: true}));
                return true;
            }""",
            value,
        )
        if not ok:
            return []
        page.wait_for_timeout(1200)
        records = _extract_visible_rows(page, max_rows=80)
        _log(logger, "verify_search_title", "ok", f"Search verification query returned {len(records)} row(s).", page=page, extra={"query": value[:120]})
        return records
    except Exception as exc:
        _log(logger, "verify_search_title", "warning", f"Search verification query failed: {exc}", page=page, extra={"query": value[:120]})
        return []


def _verify_publishing_tasks(page: Any, contexts: list[dict[str, Any]], logger: Any | None = None, state: Any | None = None) -> dict[str, Any]:
    result: dict[str, Any] = {
        "status": "running",
        "verify_result": "not_run",
        "product_count": len(contexts),
        "expected_publishing_records": len(contexts) * 2,
        "actual_matched_publishing_records": 0,
        "success_product_count": 0,
        "failed_product_count": 0,
        "products": [],
        "publishing_screenshot_path": "",
        "verification_order": [target["label"] for target in TASK_VERIFY_TARGETS],
        "verification_screenshots": {},
        "verification_errors": [],
    }

    task_entries: list[tuple[dict[str, Any], int, dict[str, Any]]] = []
    for ctx in contexts:
        idx = int(ctx.get("source_product_index", 0) or 0)
        tasks = ctx.get("publish_tasks") if isinstance(ctx.get("publish_tasks"), list) else _build_publish_tasks(ctx)
        for task in tasks[:2]:
            task["publishing_record_count"] = 0
            task["publishing_verify_status"] = "not_found"
            task["publishing_match_text"] = ""
            task["publishing_match_reasons"] = []
            task["verification_page_label"] = ""
            task["verification_page_name"] = ""
            task["verification_status"] = ""
            task["verification_url"] = ""
            task["validation_page"] = ""
            task["task_completed"] = False
            task["publish_result"] = "not_found"
            task["publish_failure_reason"] = ""
            task["failure_reason"] = "not_found_in_online_publishing_publishFail"
            task_entries.append((ctx, idx, task))
        ctx["publish_tasks"] = tasks

    used_record_keys: set[str] = set()
    product_records_by_idx: dict[int, list[dict[str, Any]]] = {}
    actual = 0
    for target in TASK_VERIFY_TARGETS:
        label = str(target["label"])
        try:
            page.goto(str(target["url"]), wait_until="domcontentloaded", timeout=12000)
            legacy._wait_ready(page)
            page.wait_for_timeout(1000)
            if label == "online":
                _click_text(page, ["所有分类"], timeout_ms=1500)
                page.wait_for_timeout(1000)
        except Exception as exc:
            result["verification_errors"].append({"target": label, "error": str(exc)})
            _log(logger, "verify_task_records", "warning", f"Could not open {target['name']}: {exc}", page=page)
            continue

        records = _extract_visible_rows(page, max_rows=180)
        screenshot_path = legacy._safe_take_screenshot(page, str(target["screenshot"]))
        result["verification_screenshots"][label] = screenshot_path
        if label == "publishing":
            result["publishing_screenshot_path"] = screenshot_path

        def find_best_for_task(ctx: dict[str, Any], task: dict[str, Any], candidate_records: list[dict[str, Any]], searched_title: bool = False) -> tuple[int, dict[str, Any], list[str], str] | None:
            scored: list[tuple[int, dict[str, Any], list[str], str]] = []
            for record in candidate_records:
                record_key = legacy._record_key(record)
                if record_key in used_record_keys:
                    continue
                score, reasons = _task_match_score(record, ctx, task)
                if searched_title and score > 0 and "searched_title" not in reasons:
                    score += 2
                    reasons = reasons + ["searched_title"]
                if _task_match_accepted(score, reasons):
                    scored.append((score, record, reasons, record_key))
            scored.sort(key=lambda item: item[0], reverse=True)
            return scored[0] if scored else None

        for ctx, idx, task in task_entries:
            if task.get("publishing_verify_status") == "matched":
                continue
            best = find_best_for_task(ctx, task, records)
            if not best:
                search_queries = [
                    str(task.get("title_clean") or "").strip(),
                    str(task.get("sku_new") or "").strip(),
                    str(task.get("title_before") or "").strip(),
                ]
                seen_queries: set[str] = set()
                for query in search_queries:
                    query = query[:180].strip()
                    if len(query) < 8 or query.lower() in seen_queries:
                        continue
                    seen_queries.add(query.lower())
                    searched_records = _search_current_list_by_text(page, query, logger=logger)
                    best = find_best_for_task(ctx, task, searched_records, searched_title=True)
                    if best:
                        break
            if not best:
                continue
            _, record, reasons, record_key = best
            used_record_keys.add(record_key)
            actual += 1
            product_records_by_idx.setdefault(idx, []).append({**record, "verification_page_label": label, "verification_page_name": target["name"]})
            task["publishing_record_count"] = 1
            task["publishing_verify_status"] = "matched"
            task["publishing_match_text"] = _record_text(record)[:1000]
            task["publishing_match_reasons"] = reasons
            task["verification_page_label"] = label
            task["verification_page_name"] = target["name"]
            task["verification_status"] = label
            task["verification_url"] = str(target["url"])
            publish_failure_reason = _extract_publish_failure_reason(record) if label == "publishFail" else ""
            task["validation_page"] = label
            task["task_completed"] = True
            task["publish_result"] = _publish_result_for_validation_page(label)
            task["publish_failure_reason"] = publish_failure_reason
            task["failure_reason"] = publish_failure_reason if label == "publishFail" else ""

    for ctx in contexts:
        idx = int(ctx.get("source_product_index", 0) or 0)
        tasks = ctx.get("publish_tasks") if isinstance(ctx.get("publish_tasks"), list) else _build_publish_tasks(ctx)
        matched_task_count = sum(1 for task in tasks[:2] if isinstance(task, dict) and task.get("publishing_verify_status") == "matched")
        product_records = product_records_by_idx.get(idx, [])
        if matched_task_count >= 2:
            failure_reason = ""
            success = True
        elif matched_task_count == 1:
            failure_reason = "only_one_task_record_found"
            success = False
        else:
            failure_reason = "no_task_record_found"
            success = False
        ctx["publish_tasks"] = tasks
        ctx["publishing_records_count"] = matched_task_count
        ctx["completion_records_count"] = matched_task_count
        ctx["distinct_publish_records_count"] = matched_task_count
        ctx["verify_result"] = "dual_publish_task_verified" if success else "dual_publish_verify_failed"
        ctx["failure_reason"] = "" if success else failure_reason
        result["success_product_count"] += 1 if success else 0
        result["failed_product_count"] += 0 if success else 1
        result["products"].append({
            "source_product_index": idx,
            "task_marker": ctx.get("task_marker", ""),
            "source_list_title": ctx.get("source_list_title", ""),
            "publishing_records_count": matched_task_count,
            "completion_records_count": matched_task_count,
            "failure_reason": failure_reason,
            "success": success,
            "publishing_records": _unique_records(product_records)[:4],
            "tasks": tasks,
        })
    result["actual_matched_publishing_records"] = actual
    all_success = bool(contexts) and result["success_product_count"] == len(contexts) and actual >= result["expected_publishing_records"]
    result["status"] = "success" if all_success else "failed"
    result["verify_result"] = "publishing_tasks_verified" if all_success else "publishing_tasks_verify_failed"
    _log(logger, "verify_publishing_tasks", result["status"], f"Task verification completed: matched={actual}/{result['expected_publishing_records']}", page=page, screenshot_path=result.get("publishing_screenshot_path", ""), extra=result)
    if state:
        state.update(dxm_twice_publishing_verify=result)
    return result


def _image_report_values(image_result: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(image_result, dict):
        return {"before": "", "after": "", "filled_to_10": "否", "shuffled": "否", "set_800px": "未执行", "child_checked": "未执行", "child_result": "unknown"}
    before = image_result.get("selected_before") or image_result.get("before_count") or image_result.get("initial_selected_count") or image_result.get("original_count") or ""
    after = image_result.get("selected_final") or image_result.get("selected_after") or image_result.get("after_count") or image_result.get("final_selected_count") or image_result.get("final_count") or ""
    try:
        filled_to_10 = "是" if int(after) >= 10 else "否"
    except Exception:
        filled_to_10 = "否"
    return {
        "before": before,
        "after": after,
        "filled_to_10": filled_to_10,
        "shuffled": "是" if image_result.get("status") == "ok" else "否",
        "set_800px": image_result.get("set_800px_status", "未执行"),
        "child_checked": image_result.get("child_check_status", "未执行"),
        "child_result": image_result.get("child_check_result", "unknown"),
        "variant_preview_action": image_result.get("variant_preview_action", ""),
        "variant_preview_before": json.dumps(image_result.get("variant_preview_before", image_result.get("variant_preview_800", {}).get("before", "")), ensure_ascii=False, default=str)[:800],
        "variant_preview_after": json.dumps(image_result.get("variant_preview_after", image_result.get("variant_preview_800", {}).get("after", "")), ensure_ascii=False, default=str)[:800],
        "variant_preview_source": image_result.get("variant_preview_source", image_result.get("variant_preview_800", {}).get("variant_preview_source", "")),
    }


def _required_counts(publish_result: dict[str, Any]) -> tuple[int, int, str]:
    if not isinstance(publish_result, dict):
        return 0, 0, ""
    ensure = publish_result.get("ensure", {}) if isinstance(publish_result.get("ensure"), dict) else {}
    errors = ensure.get("errors") if isinstance(ensure.get("errors"), list) else []
    handled = ensure.get("handled") if isinstance(ensure.get("handled"), list) else []
    unhandled = ensure.get("unhandled") if isinstance(ensure.get("unhandled"), list) else []
    scanned = len(errors) + len(handled) + len(unhandled)
    processed = len(handled)
    return scanned, processed, "; ".join(str(item)[:120] for item in unhandled)


def _write_publishing_reports(contexts: list[dict[str, Any]], verification: dict[str, Any]) -> dict[str, str]:
    reports_dir = PROJECT_ROOT / "data" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = reports_dir / f"dxm_publish_twice_result_{stamp}.json"
    xlsx_path = reports_dir / f"dxm_publish_twice_result_{stamp}.xlsx"
    verify_products = {int(item.get("source_product_index", 0) or 0): item for item in verification.get("products", []) if isinstance(item, dict)}
    rows: list[dict[str, Any]] = []
    for ctx in contexts:
        idx = int(ctx.get("source_product_index", 0) or 0)
        verify = verify_products.get(idx, {})
        tasks = ctx.get("publish_tasks") if isinstance(ctx.get("publish_tasks"), list) else _build_publish_tasks(ctx)
        old_price, new_price = _sku_price_summary(ctx.get("first_price_result", {}))
        for task in tasks[:2]:
            round_no = int(task.get("publish_round") or 0)
            image = _image_report_values(task.get("image_result", {}))
            publish_result = ctx.get("first_publish_result" if round_no == 1 else "second_publish_result", {})
            required_total, required_done, required_unhandled = _required_counts(publish_result)
            final_success = task.get("publishing_verify_status") == "matched"
            origin = publish_result.get("origin", {}) if isinstance(publish_result.get("origin"), dict) else {}
            country = origin.get("country") or "中国大陆"
            province = origin.get("province") or "广东省"
            description_image = publish_result.get("description_image", {}) if isinstance(publish_result.get("description_image"), dict) else {}
            package_info = publish_result.get("package_info", {}) if isinstance(publish_result.get("package_info"), dict) else {}
            size_weight_source = publish_result.get("size_weight_source") or "fallback_random"
            package_image_source = package_info.get("package_image_source") or publish_result.get("package_image_source") or "fallback_random"
            notes: list[str] = []
            if ctx.get("source_title_not_found_used_first_row"):
                notes.append("source_title_not_found_used_first_row")
            rows.append({
                "源产品序号": idx,
                "发布轮次": round_no,
                "task_marker": task.get("task_marker", ctx.get("task_marker", "")),
                "源商品标题": ctx.get("source_list_title", ""),
                "清理前标题": task.get("title_before", ""),
                "清理后标题": task.get("title_clean", ""),
                "命中禁词": "是" if task.get("title_forbidden_hit") else "否",
                "已清理禁词": ", ".join(task.get("cleaned_forbidden_terms") or []),
                "编辑页URL": task.get("edit_url", ""),
                "编辑ID": task.get("edit_id", ""),
                "是否quoteEdit": "是" if task.get("is_quote_edit") else "否",
                "图片原数量": image["before"],
                "图片最终数量": image["after"],
                "是否选满10张": image["filled_to_10"],
                "图片是否打乱": image["shuffled"],
                "图片是否800": image["set_800px"],
                "是否设置800px": image["set_800px"],
                "variant_preview_action": image["variant_preview_action"],
                "variant_preview_before": image["variant_preview_before"],
                "variant_preview_after": image["variant_preview_after"],
                "variant_preview_source": image["variant_preview_source"],
                "是否检测儿童图片": image["child_checked"],
                "儿童图片检测结果": image["child_result"],
                "SKU原值": task.get("sku_old", ""),
                "SKU新值": task.get("sku_new", ""),
                "SKU原价": old_price if round_no == 1 else "",
                "SKU新价": new_price if round_no == 1 else "",
                "产地": country,
                "省份": province,
                "原类目": task.get("category_old", ""),
                "新类目": task.get("category_new", ""),
                "类目": task.get("category_new", "") or task.get("category_old", ""),
                "尺寸重量来源": size_weight_source,
                "raw_dimension_text": publish_result.get("raw_dimension_text", ""),
                "length_cm": publish_result.get("length_cm", ""),
                "width_cm": publish_result.get("width_cm", ""),
                "height_cm": publish_result.get("height_cm", ""),
                "weight_g": publish_result.get("weight_g", ""),
                "外包装图片来源": package_image_source,
                "描述图片模块是否补齐": "是" if description_image.get("status") == "ok" else "否",
                "必填项扫描数量": required_total,
                "必填项已处理数量": required_done,
                "未处理必填项": required_unhandled,
                "发布提交状态": task.get("submit_status", ""),
                "验收落点": task.get("validation_page") or task.get("verification_page_label", ""),
                "任务是否完成": "是" if task.get("task_completed") or final_success else "否",
                "发布结果": task.get("publish_result") or _publish_result_for_validation_page(str(task.get("verification_page_label") or "")),
                "验收页面": task.get("verification_page_name", ""),
                "验收任务状态": task.get("verification_status", ""),
                "验收URL": task.get("verification_url", ""),
                "发布中核验状态": task.get("publishing_verify_status", ""),
                "发布中记录数": task.get("publishing_record_count", 0),
                "发布中匹配文本": task.get("publishing_match_text", ""),
                "发布失败原因": task.get("publish_failure_reason", ""),
                "失败原因": task.get("failure_reason") or verify.get("failure_reason") or ctx.get("failure_reason", ""),
                "备注": "; ".join(notes),
                "最终是否成功": "成功" if final_success else "失败",
            })
    payload = {"summary": verification, "products": contexts, "rows": rows}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    try:
        import pandas as pd

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df = pd.DataFrame(rows)
            df.to_excel(writer, index=False, sheet_name="任务验收")
            ws = writer.book["任务验收"]
            ws.freeze_panes = "A2"
            for column_cells in ws.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 60)
    except Exception:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "任务验收"
        headers = list(rows[0].keys()) if rows else ["结果"]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
    return {"json_path": str(json_path), "xlsx_path": str(xlsx_path)}


def _write_field_fill_test_report(row: dict[str, Any]) -> dict[str, str]:
    reports_dir = PROJECT_ROOT / "data" / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = reports_dir / f"dxm_field_fill_test_{stamp}.json"
    xlsx_path = reports_dir / f"dxm_field_fill_test_{stamp}.xlsx"
    payload = {"summary": row, "rows": [row]}
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    try:
        import pandas as pd

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df = pd.DataFrame([row])
            df.to_excel(writer, index=False, sheet_name="字段补全测试")
            ws = writer.book["字段补全测试"]
            ws.freeze_panes = "A2"
            for column_cells in ws.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 60)
    except Exception:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "字段补全测试"
        headers = list(row.keys())
        ws.append(headers)
        ws.append([row.get(header, "") for header in headers])
        ws.freeze_panes = "A2"
        wb.save(xlsx_path)
    return {"json_path": str(json_path), "xlsx_path": str(xlsx_path)}


def _summarize_ocr_text(dimension_info: dict[str, Any]) -> str:
    parts: list[str] = []
    for item in dimension_info.get("image_ocr_items", []) if isinstance(dimension_info.get("image_ocr_items"), list) else []:
        if not isinstance(item, dict):
            continue
        text = str(item.get("combined_text") or "").strip()
        if text:
            parts.append(f"#{item.get('index')}: {text[:500]}")
    return "\n".join(parts[:5])


def run_dxm_field_fill_test(page: Any, config: dict[str, Any], logger: Any | None = None, state: Any | None = None) -> dict[str, Any]:
    """Open one collected draft product, fill field data, and stop before publish."""
    _ensure_utf8_console()
    flow_config = config.get("dxm_field_fill_test", {}) if isinstance(config, dict) else {}
    edit_url = str(flow_config.get("edit_url") or "").strip()
    _trim_extra_browser_pages(page, max_pages=1, logger=logger)
    if edit_url:
        page.goto(edit_url, wait_until="domcontentloaded", timeout=12000)
        legacy._wait_ready(page)
        page.wait_for_timeout(700)
        edit_page = page
        source = {
            "source_index": 1,
            "title": legacy._safe_read_title(edit_page),
            "sku": legacy._read_first_sku_from_edit_page(edit_page),
            "price": "",
            "row_text_preview": "",
            "current_edit_page": True,
            "edit_url": edit_url,
        }
    else:
        page = _prefer_source_list_page(page)
        sources = _extract_source_products(page, 1, logger=logger)
        if not sources:
            screenshot_path = legacy._safe_take_screenshot(page, "dxm_field_fill_no_source")
            return {"status": "failed", "failure_reason": "no_source_product_found", "screenshot_path": screenshot_path}
        source = sources[0]
        edit_page = _open_source_edit(page, source, 1, logger=logger, state=state)
    title = legacy._safe_read_title(edit_page) or str(source.get("title") or "")
    context = {
        "source_product_index": 1,
        "source_list_title": source.get("title", ""),
        "source_list_sku": source.get("sku", ""),
        "edit_url": edit_page.url,
        "title": title,
    }

    dimension_info = extract_dimensions_from_product_images(edit_page, context)
    dimension_result = fill_variant_dimensions_and_weight_from_images(edit_page, dimension_info, logger=logger)

    fixed_config = dict(config)
    defaults = dict(fixed_config.get("product_defaults", {}) if isinstance(fixed_config.get("product_defaults"), dict) else {})
    defaults["origin_country"] = "中国大陆"
    defaults["origin_provinces"] = ["广东省"]
    fixed_config["product_defaults"] = defaults
    origin_result = select_origin_country_and_province(edit_page, fixed_config, logger=logger, state=state)

    variant_preview_result = legacy._ensure_variant_preview_images_square_800(edit_page, stage="field_fill_test", logger=logger, add_missing=True)
    package_result = choose_package_image_with_dimension_priority(edit_page, {**context, **dimension_info})
    screenshot_path = legacy._safe_take_screenshot(edit_page, "dxm_field_fill_test_result")
    row = {
        "编辑页URL": edit_page.url,
        "源商品标题": title,
        "带尺寸图片序号": dimension_info.get("dimension_candidate_index", ""),
        "图片识别文字": _summarize_ocr_text(dimension_info),
        "识别到的尺寸原文": dimension_info.get("raw_dimension_text", ""),
        "填入长cm": dimension_info.get("length_cm", ""),
        "填入宽cm": dimension_info.get("width_cm", ""),
        "填入高cm": dimension_info.get("height_cm", ""),
        "填入重量g": dimension_info.get("weight_g", ""),
        "尺寸重量来源": dimension_info.get("source", ""),
        "重量来源": "fallback" if "weight_fallback" in str(dimension_info.get("source", "")) or dimension_info.get("source") == "fallback_random" else "image_detected",
        "variant_preview_action": variant_preview_result.get("variant_preview_action", ""),
        "variant_preview_before": json.dumps(variant_preview_result.get("variant_preview_before", []), ensure_ascii=False, default=str)[:800],
        "variant_preview_after": json.dumps(variant_preview_result.get("variant_preview_after", []), ensure_ascii=False, default=str)[:800],
        "variant_preview_source": variant_preview_result.get("variant_preview_source", ""),
        "外包装图片序号": package_result.get("selected_index", ""),
        "外包装图片是否带尺寸": "是" if package_result.get("has_dimension_mark") else "否",
        "外包装图片来源": package_result.get("package_image_source", ""),
        "产地": origin_result.get("country", "中国大陆"),
        "省份": origin_result.get("province", "广东省"),
        "是否点击发布": "否",
        "备注": "; ".join([
            f"edit_url={edit_page.url}",
            f"dimension_status={dimension_result.get('status')}",
            f"package_status={package_result.get('status')}",
            f"origin_status={origin_result.get('status')}",
            f"variant_preview_action={variant_preview_result.get('variant_preview_action')}",
            f"ocr_debug={dimension_info.get('image_ocr_debug_path', '')}",
            f"package_text={str(package_result.get('selected_image_text', ''))[:200]}",
        ]),
        "截图路径": screenshot_path,
    }
    reports = _write_field_fill_test_report(row)
    result = {
        "status": "success" if dimension_result.get("status") == "ok" else "manual_required",
        "source_product_count": 1,
        "source": source,
        "title": title,
        "dimension_info": dimension_info,
        "dimension_result": dimension_result,
        "origin_result": origin_result,
        "variant_preview_result": variant_preview_result,
        "package_result": package_result,
        "screenshot_path": screenshot_path,
        "reports": reports,
        "clicked_publish": False,
    }
    _log(logger, "dxm_field_fill_test", result["status"], "Field fill test completed without clicking publish.", page=edit_page, screenshot_path=screenshot_path, extra=result)
    if state:
        state.update(dxm_field_fill_test=result)
    return result


def run_dxm_publish_twice(page: Any, config: dict[str, Any], logger: Any | None = None, state: Any | None = None) -> dict[str, Any]:
    _ensure_utf8_console()
    _patch_legacy_short_title_cleaner()
    flow_config = config.get("dxm_publish_twice", {}) if isinstance(config, dict) else {}
    product_count = int(flow_config.get("source_product_count", flow_config.get("product_count", DEFAULT_PRODUCT_COUNT)))
    product_count = max(1, product_count)
    target_title = str(flow_config.get("target_title") or "").strip()
    _trim_extra_browser_pages(page, max_pages=1, logger=logger)
    page = _prefer_source_list_page(page)
    source_list_url = DRAFT_PRODUCTS_URL
    sources = _extract_source_products(page, product_count, logger=logger, target_title=target_title)
    result: dict[str, Any] = {
        "status": "running",
        "source_product_count": product_count,
        "product_count": product_count,
        "selected_source_count": len(sources),
        "expected_publishing_records": product_count * 2,
        "source_list_url": source_list_url,
        "products": [],
        "reports": {},
    }
    if len(sources) < product_count:
        result["not_enough_items"] = True
        result["not_enough_items_message"] = f"Requested {product_count} products but only found {len(sources)} source item(s)."
    if not sources:
        screenshot_path = legacy._safe_take_screenshot(page, "dxm_twice_no_source_products")
        result.update({"status": "failed", "failure_reason": "not_enough_items", "screenshot_path": screenshot_path})
        return result

    current_page = page
    contexts: list[dict[str, Any]] = []
    for idx, source in enumerate(sources, start=1):
        if source_list_url:
            current_page = _return_to_source_list(current_page, source_list_url, logger=logger)
        try:
            current_page, context = _publish_source_twice(current_page, source, idx, config, logger=logger, state=state)
            contexts.append(context)
            result["products"].append({"source_product_index": idx, "status": "submitted", "context": context})
        except legacy.DxmTwiceFlowError as exc:
            failed_context = {
                "source_product_index": idx,
                "source_list_title": source.get("title", ""),
                "source_list_sku": source.get("sku", ""),
                "failure_reason": exc.step,
                "failure_message": exc.message,
                "screenshot_path": exc.screenshot_path,
                "verify_result": "dual_publish_verify_failed",
            }
            contexts.append(failed_context)
            result["products"].append({"source_product_index": idx, "status": "failed", "context": failed_context})
            _log(logger, "dxm_publish_twice_product", "failed", exc.message, page=current_page, screenshot_path=exc.screenshot_path)
        except Exception as exc:
            screenshot_path = legacy._safe_take_screenshot(current_page, f"dxm_twice_product_{idx}_error")
            failed_context = {
                "source_product_index": idx,
                "source_list_title": source.get("title", ""),
                "source_list_sku": source.get("sku", ""),
                "failure_reason": "unexpected_error",
                "failure_message": str(exc),
                "screenshot_path": screenshot_path,
                "verify_result": "dual_publish_verify_failed",
            }
            contexts.append(failed_context)
            result["products"].append({"source_product_index": idx, "status": "failed", "context": failed_context})
            _log(logger, "dxm_publish_twice_product", "failed", f"Unexpected product error: {exc}", page=current_page, screenshot_path=screenshot_path)

    verification = _verify_publishing_tasks(current_page, contexts, logger=logger, state=state)
    retried_indices: list[int] = []
    failed_verify_indices = [
        int(item.get("source_product_index", 0) or 0)
        for item in verification.get("products", [])
        if isinstance(item, dict) and not item.get("success")
    ]
    retry_on_verify_failure = bool(flow_config.get("retry_on_verify_failure", False))
    if not retry_on_verify_failure:
        failed_verify_indices = []
    for failed_idx in failed_verify_indices:
        if failed_idx < 1 or failed_idx > len(sources):
            continue
        retried_indices.append(failed_idx)
        source = dict(sources[failed_idx - 1])
        previous_context = next((ctx for ctx in contexts if int(ctx.get("source_product_index", 0) or 0) == failed_idx), {})
        if previous_context.get("first_edit_url"):
            source["edit_url"] = previous_context.get("first_edit_url")
            source["title"] = previous_context.get("source_list_title") or previous_context.get("first_title_before") or source.get("title", "")
            source["sku"] = previous_context.get("source_list_sku") or source.get("sku", "")
        try:
            current_page = _return_to_source_list(current_page, DRAFT_PRODUCTS_URL, logger=logger)
            current_page, retry_context = _publish_source_twice(current_page, source, failed_idx, config, logger=logger, state=state)
            retry_context["retry_attempt"] = 1
            previous_tasks = previous_context.get("publish_tasks") if isinstance(previous_context.get("publish_tasks"), list) else []
            previous_matched = [dict(task) for task in previous_tasks if isinstance(task, dict) and task.get("publishing_verify_status") == "matched"]
            retry_tasks = retry_context.get("publish_tasks") if isinstance(retry_context.get("publish_tasks"), list) else _build_publish_tasks(retry_context)
            if previous_matched and len(retry_tasks) >= 2:
                first_actual = previous_matched[0]
                second_actual = dict(retry_tasks[1])
                first_actual["publish_round"] = 1
                second_actual["publish_round"] = 2
                retry_context["publish_tasks"] = [first_actual, second_actual]
                retry_context["merged_previous_attempt_record"] = True
                retry_context["previous_attempt_task_marker"] = previous_context.get("task_marker", "")
                retry_context["source_list_title"] = previous_context.get("source_list_title") or retry_context.get("source_list_title", "")
                retry_context["source_list_sku"] = previous_context.get("source_list_sku") or retry_context.get("source_list_sku", "")
                _log(logger, "dxm_publish_twice_retry_merge", "ok", "Merged the previously verified quoteEdit record with the retry quoteEdit task.", page=current_page)
            replace_at = next((pos for pos, ctx in enumerate(contexts) if int(ctx.get("source_product_index", 0) or 0) == failed_idx), None)
            if replace_at is None:
                contexts.append(retry_context)
            else:
                contexts[replace_at] = retry_context
            _log(logger, "dxm_publish_twice_retry", "ok", f"Retried source product {failed_idx}.", page=current_page)
        except legacy.DxmTwiceFlowError as exc:
            replace_at = next((pos for pos, ctx in enumerate(contexts) if int(ctx.get("source_product_index", 0) or 0) == failed_idx), None)
            if replace_at is not None:
                contexts[replace_at]["retry_attempt"] = 1
                contexts[replace_at]["failure_reason"] = "source_task_not_found_in_draft_after_failure" if exc.step == "open_source_product_edit" else exc.step
                contexts[replace_at]["failure_message"] = exc.message
                contexts[replace_at]["screenshot_path"] = exc.screenshot_path
            _log(logger, "dxm_publish_twice_retry", "failed", exc.message, page=current_page, screenshot_path=exc.screenshot_path)
        except Exception as exc:
            screenshot_path = legacy._safe_take_screenshot(current_page, f"dxm_twice_retry_{failed_idx}_error")
            replace_at = next((pos for pos, ctx in enumerate(contexts) if int(ctx.get("source_product_index", 0) or 0) == failed_idx), None)
            if replace_at is not None:
                contexts[replace_at]["retry_attempt"] = 1
                contexts[replace_at]["failure_reason"] = "retry_unexpected_error"
                contexts[replace_at]["failure_message"] = str(exc)
                contexts[replace_at]["screenshot_path"] = screenshot_path
            _log(logger, "dxm_publish_twice_retry", "failed", f"Retry source product {failed_idx} failed: {exc}", page=current_page, screenshot_path=screenshot_path)
    if retried_indices:
        verification = _verify_publishing_tasks(current_page, contexts, logger=logger, state=state)
        result["retried_source_indices"] = retried_indices
    result["verification"] = verification
    result["actual_matched_publishing_records"] = verification.get("actual_matched_publishing_records", 0)
    result["success_product_count"] = verification.get("success_product_count", 0)
    result["failed_product_count"] = verification.get("failed_product_count", 0)
    result["verify_result"] = verification.get("verify_result", "not_run")
    result["status"] = "success" if verification.get("verify_result") == "publishing_tasks_verified" and len(contexts) == product_count else "publishing_tasks_verify_failed"
    result["reports"] = _write_publishing_reports(contexts, verification)
    xlsx_path = str(result.get("reports", {}).get("xlsx_path") or "")
    if xlsx_path:
        try:
            os.startfile(xlsx_path)  # type: ignore[attr-defined]
            result["excel_opened"] = True
        except Exception as exc:
            result["excel_opened"] = False
            result["excel_open_error"] = str(exc)
    if state:
        state.update(dxm_publish_twice=result)
    _log(logger, "dxm_publish_twice_done", result["status"], f"DXM publish twice source_product_count={product_count}; actual_matched_publishing_records={result.get('actual_matched_publishing_records')}", page=current_page, extra={"reports": result["reports"]})
    return result
